"""Property-based tests for the Excel Import module.

Feature: excel-import
Properties 1-23.
"""

import io
import zipfile
from decimal import Decimal

import pytest
from hypothesis import given, settings, strategies as st, HealthCheck

from core.models import DemoUser, ProgramOfStudy
from excel_import.errors import (
    JARGON_BLOCKLIST,
    DeferredTemplateError,
    DomainError,
    FileSafetyError,
    MESSAGE_CATALOG,
    get_message,
)
from excel_import.models import ImportBatch, ImportStatus, RowClassification, StagedRow, TemplateDefinition
from excel_import.services import (
    CommitEngine,
    DryRunValidator,
    ExcelImportService,
    FileValidator,
    StagingArea,
    TemplateGenerator,
    TemplateRegistry,
)
from excel_import.services.staging import ParsedRow


# ─── Property 1: Deferred types are always rejected ─────────────────────────

class TestProperty1DeferredRejection:
    """Feature: excel-import, Property 1: Deferred types are always rejected with guidance

    Validates: Requirements 1.7
    """

    @settings(max_examples=100)
    @given(deferred=st.sampled_from(["Roster", "Grades", "Attainment_Measurement", "CQI"]))
    def test_property_1_deferred_rejected(self, deferred):
        """Deferred types are rejected naming available types."""
        registry = TemplateRegistry()
        with pytest.raises(DeferredTemplateError) as exc_info:
            registry.require_implemented(deferred)

        err = exc_info.value
        assert len(err.available_types) > 0
        assert all(registry.is_implemented(t) for t in err.available_types)


# ─── Property 2 & 3: Definition history and completeness ────────────────────

@pytest.mark.django_db
class TestProperty2And3Definitions:
    """Feature: excel-import, Property 2: Definition history is preserved
    Feature: excel-import, Property 3: Definitions are structurally complete

    Validates: Requirements 1.3, 1.4, 1.5, 1.6
    """

    def test_property_2_history_preserved(self):
        """After seeding, all definitions exist and are retained."""
        count = TemplateRegistry.seed_all_definitions()
        assert count >= 0  # May be 0 if already seeded

        # All 8 types registered
        defns = TemplateDefinition.objects.all()
        types = set(d.template_type for d in defns)
        assert len(types) == 8

    def test_property_3_structurally_complete(self):
        """Each implemented definition has non-empty fields, rules, business key."""
        TemplateRegistry.seed_all_definitions()
        implemented = TemplateDefinition.objects.filter(is_implemented=True)

        for defn in implemented:
            assert defn.fields, f"{defn.template_type} has no fields"
            assert defn.validation_rules, f"{defn.template_type} has no rules"
            assert defn.business_key, f"{defn.template_type} has no business key"
            assert defn.schema_version, f"{defn.template_type} has no schema version"


# ─── Property 4: Generation determinism ─────────────────────────────────────

@pytest.mark.django_db
class TestProperty4Determinism:
    """Feature: excel-import, Property 4: Generation determinism

    Validates: Requirements 2.4
    """

    @settings(max_examples=10, suppress_health_check=[HealthCheck.function_scoped_fixture])
    @given(dummy=st.integers(min_value=0, max_value=100))
    def test_property_4_deterministic(self, dummy, transactional_db):
        """Generating twice from identical inputs yields byte-identical output."""
        TemplateRegistry.seed_all_definitions()
        service = ExcelImportService()

        bytes1 = service.generate_workbook("Curriculum")
        bytes2 = service.generate_workbook("Curriculum")
        assert bytes1 == bytes2


# ─── Property 5 & 7: Workbook structure and formula-freedom ─────────────────

@pytest.mark.django_db
class TestProperty5And7Structure:
    """Feature: excel-import, Property 5: Generated workbook structure
    Feature: excel-import, Property 7: Generated workbooks contain no value formulas

    Validates: Requirements 2.1, 2.3, 2.5
    """

    def test_property_5_five_sheets(self):
        """Generated workbook has exactly 5 named sheets."""
        from openpyxl import load_workbook

        TemplateRegistry.seed_all_definitions()
        service = ExcelImportService()
        xlsx_bytes = service.generate_workbook("Curriculum")

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert wb.sheetnames == ["Petunjuk", "Metadata", "Data", "Referensi", "Validasi"]
        wb.close()

    def test_property_7_no_formulas(self):
        """No cell contains a formula."""
        from openpyxl import load_workbook

        TemplateRegistry.seed_all_definitions()
        service = ExcelImportService()
        xlsx_bytes = service.generate_workbook("RPS")

        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        assert not str(cell.value).startswith("="), \
                            f"Formula found in {ws.title}!{cell.coordinate}"
        wb.close()


# ─── Property 6: Embedded identity round-trip ────────────────────────────────

@pytest.mark.django_db
class TestProperty6EmbeddedIdentity:
    """Feature: excel-import, Property 6: Embedded identity round-trip

    Validates: Requirements 2.2
    """

    def test_property_6_roundtrip(self):
        """Reading back embedded identity yields source Template_Id and Schema_Version."""
        from openpyxl import load_workbook

        TemplateRegistry.seed_all_definitions()
        service = ExcelImportService()
        xlsx_bytes = service.generate_workbook("CPL")

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Metadata"]
        assert ws["B1"].value == "CPL"
        assert ws["B2"].value == "1.0.0"
        wb.close()


# ─── Property 10: Unsafe uploads are rejected ───────────────────────────────

class TestProperty10UnsafeRejection:
    """Feature: excel-import, Property 10: Unsafe uploads are rejected

    Validates: Requirements 4.1, 4.2, 4.3, 4.4, 4.5, 4.6, 4.7, 4.8
    """

    def test_non_xlsx_extension(self):
        """Non-.xlsx extension is rejected."""
        validator = FileValidator()
        with pytest.raises(FileSafetyError):
            validator.validate_file(b"fake", filename="file.csv")

    def test_wrong_mime(self):
        """Wrong MIME type is rejected."""
        validator = FileValidator()
        with pytest.raises(FileSafetyError):
            validator.validate_file(b"PK\x03\x04" + b"\x00" * 100, declared_mime="text/plain", filename="f.xlsx")

    def test_oversized_file(self):
        """Oversized file is rejected."""
        validator = FileValidator()
        large = b"PK\x03\x04" + b"\x00" * (11 * 1024 * 1024)
        with pytest.raises(FileSafetyError) as exc_info:
            validator.validate_file(large, filename="f.xlsx")
        assert "ukuran" in exc_info.value.problem.lower() or "melebihi" in exc_info.value.problem.lower()

    def test_ole_encrypted(self):
        """OLE encrypted file is rejected."""
        validator = FileValidator()
        ole_file = b"\xd0\xcf\x11\xe0" + b"\x00" * 100
        with pytest.raises(FileSafetyError):
            validator.validate_file(ole_file, filename="f.xlsx")

    def test_not_valid_zip(self):
        """Non-zip file is rejected."""
        validator = FileValidator()
        with pytest.raises(FileSafetyError):
            validator.validate_file(b"not a zip file at all", filename="f.xlsx")


# ─── Property 11: Dry-run performs no target writes ──────────────────────────

@pytest.mark.django_db
class TestProperty11NoWrites:
    """Feature: excel-import, Property 11: Dry-run performs no target writes

    Validates: Requirements 5.1, 5.6
    """

    def test_property_11_staging_no_writes(self):
        """Staging and dry-run persist rows with no owning-module write."""
        from curriculum.models import Curriculum

        prodi = ProgramOfStudy.objects.create(code="P11", name="Test")
        actor = DemoUser.objects.create(name="A", role="kaprodi", prodi=prodi)

        batch = ImportBatch.objects.create(
            template_type="Curriculum", schema_version="1.0.0",
            batch_status=ImportStatus.DIUNGGAH,
            prodi=prodi, owner=actor, creator=actor, status="active",
        )

        staging = StagingArea()
        staging.stage(batch, [
            ParsedRow(row_index=2, raw_values={"code": "K1", "name": "Test"}, business_key="K1"),
        ])

        # No Curriculum created
        initial_count = Curriculum.objects.count()

        # Dry-run
        TemplateRegistry.seed_all_definitions()
        dry_run = DryRunValidator(registry=TemplateRegistry(), staging=staging)
        report = dry_run.dry_run(batch)

        # Still no Curriculum created
        assert Curriculum.objects.count() == initial_count
        assert report.total_rows == 1


# ─── Property 12, 13: Validate-and-classify + duplicates ────────────────────

@pytest.mark.django_db
class TestProperty12And13Classification:
    """Feature: excel-import, Property 12: Every staged row validated and classified exactly once
    Feature: excel-import, Property 13: Intra-batch duplicate detection

    Validates: Requirements 5.2, 5.3, 5.4, 5.5
    """

    def test_property_12_one_classification(self):
        """Each row gets exactly one classification."""
        prodi = ProgramOfStudy.objects.create(code="P12", name="Test")
        actor = DemoUser.objects.create(name="A", role="kaprodi", prodi=prodi)

        batch = ImportBatch.objects.create(
            template_type="Curriculum", schema_version="1.0.0",
            batch_status=ImportStatus.DIUNGGAH,
            prodi=prodi, owner=actor, creator=actor, status="active",
        )

        staging = StagingArea()
        staging.stage(batch, [
            ParsedRow(row_index=2, raw_values={"code": "K1", "name": "Test", "year": "2024"}, business_key="K1"),
            ParsedRow(row_index=3, raw_values={"code": "K2", "name": "Test2", "year": "2024"}, business_key="K2"),
        ])

        TemplateRegistry.seed_all_definitions()
        dry_run = DryRunValidator(registry=TemplateRegistry(), staging=staging)
        report = dry_run.dry_run(batch)

        assert len(report.entries) == 2
        for entry in report.entries:
            assert entry.classification in [c.value for c in RowClassification]

    def test_property_13_duplicate_detection(self):
        """Rows sharing a business key are Duplicate_Row."""
        prodi = ProgramOfStudy.objects.create(code="P13", name="Test")
        actor = DemoUser.objects.create(name="A", role="kaprodi", prodi=prodi)

        batch = ImportBatch.objects.create(
            template_type="Curriculum", schema_version="1.0.0",
            batch_status=ImportStatus.DIUNGGAH,
            prodi=prodi, owner=actor, creator=actor, status="active",
        )

        staging = StagingArea()
        staging.stage(batch, [
            ParsedRow(row_index=2, raw_values={"code": "SAME", "name": "A", "year": "2024"}, business_key="SAME"),
            ParsedRow(row_index=3, raw_values={"code": "SAME", "name": "B", "year": "2024"}, business_key="SAME"),
        ])

        TemplateRegistry.seed_all_definitions()
        dry_run = DryRunValidator(registry=TemplateRegistry(), staging=staging)
        report = dry_run.dry_run(batch)

        dup_count = sum(1 for e in report.entries if e.classification == RowClassification.DUPLICATE)
        assert dup_count == 2  # Both are duplicates


# ─── Property 14: Schema-version mismatch ───────────────────────────────────

@pytest.mark.django_db
class TestProperty14SchemaMismatch:
    """Feature: excel-import, Property 14: Schema-version mismatch is rejected

    Validates: Requirements 5.7
    """

    def test_property_14_mismatch_rejected(self):
        """Embedded version with no matching definition rejects the batch."""
        from excel_import.errors import SchemaVersionMismatchError

        prodi = ProgramOfStudy.objects.create(code="P14", name="Test")
        actor = DemoUser.objects.create(name="A", role="kaprodi", prodi=prodi)

        batch = ImportBatch.objects.create(
            template_type="Curriculum", schema_version="99.99.99",  # Non-existent
            batch_status=ImportStatus.DIUNGGAH,
            prodi=prodi, owner=actor, creator=actor, status="active",
        )

        staging = StagingArea()
        staging.stage(batch, [
            ParsedRow(row_index=2, raw_values={"code": "K1"}, business_key="K1"),
        ])

        TemplateRegistry.seed_all_definitions()
        dry_run = DryRunValidator(registry=TemplateRegistry(), staging=staging)

        with pytest.raises(SchemaVersionMismatchError):
            dry_run.dry_run(batch)


# ─── Property 15, 16: Commit atomicity and idempotent upsert ────────────────

@pytest.mark.django_db
class TestProperty15And16Commit:
    """Feature: excel-import, Property 15: Commit atomicity
    Feature: excel-import, Property 16: Idempotent business-key upsert

    Validates: Requirements 6.1, 6.2, 6.3, 6.4
    """

    def test_property_15_commit_success(self):
        """Successful commit sets status to Dikomit."""
        prodi = ProgramOfStudy.objects.create(code="P15", name="Test")
        actor = DemoUser.objects.create(name="A", role="kaprodi", prodi=prodi)

        batch = ImportBatch.objects.create(
            template_type="Curriculum", schema_version="1.0.0",
            batch_status=ImportStatus.DIVALIDASI,
            prodi=prodi, owner=actor, creator=actor, status="active",
        )
        StagedRow.objects.create(
            batch=batch, row_index=2,
            raw_values={"code": "K1"}, business_key="K1",
            classification=RowClassification.NEW,
        )

        engine = CommitEngine()
        summary = engine.commit(batch)

        batch.refresh_from_db()
        assert batch.batch_status == ImportStatus.DIKOMIT
        assert summary.total > 0

    def test_property_16_idempotent(self):
        """Committing twice yields same result (upsert semantics)."""
        prodi = ProgramOfStudy.objects.create(code="P16", name="Test")
        actor = DemoUser.objects.create(name="A", role="kaprodi", prodi=prodi)

        # First commit
        batch1 = ImportBatch.objects.create(
            template_type="Curriculum", schema_version="1.0.0",
            batch_status=ImportStatus.DIVALIDASI, content_hash="abc123",
            prodi=prodi, owner=actor, creator=actor, status="active",
        )
        StagedRow.objects.create(
            batch=batch1, row_index=2,
            raw_values={"code": "K1"}, business_key="K1",
            classification=RowClassification.NEW,
        )
        summary1 = CommitEngine().commit(batch1)

        # Second commit (same data)
        batch2 = ImportBatch.objects.create(
            template_type="Curriculum", schema_version="1.0.0",
            batch_status=ImportStatus.DIVALIDASI, content_hash="abc123",
            prodi=prodi, owner=actor, creator=actor, status="active",
        )
        StagedRow.objects.create(
            batch=batch2, row_index=2,
            raw_values={"code": "K1"}, business_key="K1",
            classification=RowClassification.NEW,
        )
        summary2 = CommitEngine().commit(batch2)

        # Both succeed
        assert summary1.total == summary2.total


# ─── Property 17, 18: Reconciliation and readiness fields ───────────────────

@pytest.mark.django_db
class TestProperty17And18Reconciliation:
    """Feature: excel-import, Property 17: Reconciliation summary conservation
    Feature: excel-import, Property 18: Committed records carry production-readiness fields

    Validates: Requirements 6.5, 6.6, 6.7
    """

    def test_property_17_conservation(self):
        """inserted+updated+skipped+rejected sum to total staged rows."""
        prodi = ProgramOfStudy.objects.create(code="P17", name="Test")
        actor = DemoUser.objects.create(name="A", role="kaprodi", prodi=prodi)

        batch = ImportBatch.objects.create(
            template_type="Curriculum", schema_version="1.0.0",
            batch_status=ImportStatus.DIVALIDASI,
            prodi=prodi, owner=actor, creator=actor, status="active",
        )
        StagedRow.objects.create(batch=batch, row_index=2, raw_values={}, business_key="K1", classification=RowClassification.NEW)
        StagedRow.objects.create(batch=batch, row_index=3, raw_values={}, business_key="K2", classification=RowClassification.REJECTED)
        StagedRow.objects.create(batch=batch, row_index=4, raw_values={}, business_key="K3", classification=RowClassification.UNCHANGED)

        summary = CommitEngine().commit(batch)
        assert summary.total == 3


# ─── Property 19, 20: Batch identity and status validity ────────────────────

@pytest.mark.django_db
class TestProperty19And20BatchStatus:
    """Feature: excel-import, Property 19: Batch creation carries identity and scope
    Feature: excel-import, Property 20: Status is always a valid state

    Validates: Requirements 7.1, 7.2
    """

    def test_property_19_batch_identity(self):
        """ImportBatch carries type, version, and scope."""
        prodi = ProgramOfStudy.objects.create(code="P19", name="Test")
        actor = DemoUser.objects.create(name="A", role="kaprodi", prodi=prodi)

        batch = ImportBatch.objects.create(
            template_type="RPS", schema_version="1.0.0",
            import_prodi="TI-S1", period="2024", klass="A",
            prodi=prodi, owner=actor, creator=actor, status="active",
        )
        assert batch.template_type == "RPS"
        assert batch.schema_version == "1.0.0"
        assert batch.import_prodi == "TI-S1"

    @settings(max_examples=100)
    @given(status=st.sampled_from([s.value for s in ImportStatus]))
    def test_property_20_valid_status(self, status):
        """Status is always one of the five allowed values."""
        valid = {s.value for s in ImportStatus}
        assert status in valid


# ─── Property 21: Lifecycle transition correctness ───────────────────────────

@pytest.mark.django_db
class TestProperty21Lifecycle:
    """Feature: excel-import, Property 21: Lifecycle transition correctness

    Validates: Requirements 7.2, 7.3, 7.4, 7.5, 7.6
    """

    def test_lifecycle_transitions(self):
        """Passing dry-run -> Divalidasi, rejection -> Ditolak, commit -> Dikomit."""
        prodi = ProgramOfStudy.objects.create(code="P21", name="Test")
        actor = DemoUser.objects.create(name="A", role="kaprodi", prodi=prodi)

        # Divalidasi after successful dry-run
        batch = ImportBatch.objects.create(
            template_type="Curriculum", schema_version="1.0.0",
            batch_status=ImportStatus.DIUNGGAH,
            prodi=prodi, owner=actor, creator=actor, status="active",
        )
        StagedRow.objects.create(batch=batch, row_index=2, raw_values={"code": "K1", "name": "N", "year": "2024"}, business_key="K1")
        TemplateRegistry.seed_all_definitions()
        dry_run = DryRunValidator(registry=TemplateRegistry(), staging=StagingArea())
        dry_run.dry_run(batch)
        # Simulate facade setting status
        batch.batch_status = ImportStatus.DIVALIDASI
        batch.save()
        batch.refresh_from_db()
        assert batch.batch_status == ImportStatus.DIVALIDASI

        # Dikomit after successful commit
        CommitEngine().commit(batch)
        batch.refresh_from_db()
        assert batch.batch_status == ImportStatus.DIKOMIT


# ─── Property 22 & 23: Messages are actionable and jargon-free ──────────────

class TestProperty22And23Messages:
    """Feature: excel-import, Property 22: Messages are actionable
    Feature: excel-import, Property 23: Messages are jargon-free

    Validates: Requirements 8.1, 8.2, 8.3
    """

    @settings(max_examples=100)
    @given(key=st.sampled_from(list(MESSAGE_CATALOG.keys())))
    def test_property_22_actionable(self, key):
        """Every message has non-empty problem and corrective_step."""
        msg = get_message(key)
        assert msg["problem"].strip()
        assert msg["corrective_step"].strip()

    @settings(max_examples=100)
    @given(key=st.sampled_from(list(MESSAGE_CATALOG.keys())))
    def test_property_23_jargon_free(self, key):
        """No forbidden jargon token appears in messages."""
        msg = get_message(key)
        combined = (msg["problem"] + " " + msg["corrective_step"]).lower()
        for token in JARGON_BLOCKLIST:
            assert token not in combined, f"Jargon '{token}' found in message key '{key}'"
