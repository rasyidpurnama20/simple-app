"""FileValidator — layered safety pipeline for uploaded .xlsx files.

Runs ordered checks before any business parse (Requirements 4.1-4.8).
"""

from __future__ import annotations

import io
import zipfile

from excel_import.errors import FileSafetyError, get_message

# Maximum file size (10 MB)
MAX_FILE_SIZE = 10 * 1024 * 1024

# Maximum decompressed size (100 MB)
MAX_DECOMPRESSED_SIZE = 100 * 1024 * 1024

# Maximum compression ratio per entry
MAX_COMPRESSION_RATIO = 100

# OOXML MIME type
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Zip magic bytes
ZIP_MAGIC = b"PK\x03\x04"

# OLE CFB magic (encrypted files)
OLE_MAGIC = b"\xd0\xcf\x11\xe0"


class FileValidator:
    """Layered safety pipeline for uploaded Excel files."""

    def validate_file(self, file_bytes: bytes, declared_mime: str = "", filename: str = "") -> None:
        """Run the full validation pipeline. Raises FileSafetyError on failure.

        Pipeline order:
        1. Extension .xlsx
        2. MIME match
        3. Max size
        4. Valid OOXML zip
        5. Zip-bomb guard
        6. Macro detection
        7. Encryption/password
        8. External links / embedded objects
        9. Value-formula scan
        """
        self._check_extension(filename)
        self._check_mime(declared_mime)
        self._check_size(file_bytes)
        self._check_valid_zip(file_bytes)
        self._check_zip_bomb(file_bytes)
        self._check_macros(file_bytes)
        self._check_encryption(file_bytes)
        self._check_external_links(file_bytes)
        self._check_formulas(file_bytes)

    def _check_extension(self, filename: str) -> None:
        """Check .xlsx extension (Requirement 4.1)."""
        if filename and not filename.lower().endswith(".xlsx"):
            msg = get_message("extension_invalid")
            raise FileSafetyError(
                problem=msg["problem"],
                corrective_step=msg["corrective_step"],
                check_name="extension",
            )
        # Also check magic bytes will be done in _check_valid_zip

    def _check_mime(self, declared_mime: str) -> None:
        """Check MIME type (Requirement 4.2)."""
        if declared_mime and declared_mime != XLSX_MIME:
            msg = get_message("mime_mismatch")
            raise FileSafetyError(
                problem=msg["problem"],
                corrective_step=msg["corrective_step"],
                check_name="mime",
            )

    def _check_size(self, file_bytes: bytes) -> None:
        """Check file size (Requirement 4.3)."""
        if len(file_bytes) > MAX_FILE_SIZE:
            msg = get_message("size_exceeded")
            raise FileSafetyError(
                problem=msg["problem"],
                corrective_step=msg["corrective_step"],
                check_name="size",
            )

    def _check_valid_zip(self, file_bytes: bytes) -> None:
        """Check valid OOXML zip (Requirement 4.1 continued)."""
        # Check for OLE magic first (encrypted file masquerading)
        if file_bytes[:4] == OLE_MAGIC:
            msg = get_message("encrypted_file")
            raise FileSafetyError(
                problem=msg["problem"],
                corrective_step=msg["corrective_step"],
                check_name="encryption",
            )

        if file_bytes[:4] != ZIP_MAGIC:
            msg = get_message("invalid_zip")
            raise FileSafetyError(
                problem=msg["problem"],
                corrective_step=msg["corrective_step"],
                check_name="valid_zip",
            )

        try:
            with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as zf:
                # Basic validity check
                if zf.testzip() is not None:
                    msg = get_message("invalid_zip")
                    raise FileSafetyError(
                        problem=msg["problem"],
                        corrective_step=msg["corrective_step"],
                        check_name="valid_zip",
                    )
        except zipfile.BadZipFile:
            msg = get_message("invalid_zip")
            raise FileSafetyError(
                problem=msg["problem"],
                corrective_step=msg["corrective_step"],
                check_name="valid_zip",
            )

    def _check_zip_bomb(self, file_bytes: bytes) -> None:
        """Check for zip bombs (Requirement 4.4)."""
        try:
            with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as zf:
                total_decompressed = sum(info.file_size for info in zf.infolist())
                if total_decompressed > MAX_DECOMPRESSED_SIZE:
                    msg = get_message("zip_bomb")
                    raise FileSafetyError(
                        problem=msg["problem"],
                        corrective_step=msg["corrective_step"],
                        check_name="zip_bomb",
                    )
                # Check ratio per entry
                for info in zf.infolist():
                    if info.compress_size > 0:
                        ratio = info.file_size / info.compress_size
                        if ratio > MAX_COMPRESSION_RATIO:
                            msg = get_message("zip_bomb")
                            raise FileSafetyError(
                                problem=msg["problem"],
                                corrective_step=msg["corrective_step"],
                                check_name="zip_bomb",
                            )
        except zipfile.BadZipFile:
            pass  # Already caught in previous check

    def _check_macros(self, file_bytes: bytes) -> None:
        """Check for macros (Requirement 4.5)."""
        try:
            with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as zf:
                namelist = zf.namelist()
                macro_indicators = [
                    n for n in namelist
                    if "vbaProject.bin" in n or n.endswith(".bin") and "vba" in n.lower()
                ]
                if macro_indicators:
                    msg = get_message("macro_detected")
                    raise FileSafetyError(
                        problem=msg["problem"],
                        corrective_step=msg["corrective_step"],
                        check_name="macro",
                    )
        except zipfile.BadZipFile:
            pass

    def _check_encryption(self, file_bytes: bytes) -> None:
        """Check for encryption/password (Requirement 4.6)."""
        # OLE check already done in _check_valid_zip
        try:
            with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as zf:
                namelist = zf.namelist()
                encryption_indicators = [
                    n for n in namelist
                    if "EncryptedPackage" in n or "EncryptionInfo" in n
                ]
                if encryption_indicators:
                    msg = get_message("encrypted_file")
                    raise FileSafetyError(
                        problem=msg["problem"],
                        corrective_step=msg["corrective_step"],
                        check_name="encryption",
                    )
        except zipfile.BadZipFile:
            pass

    def _check_external_links(self, file_bytes: bytes) -> None:
        """Check for external links / embedded objects (Requirement 4.7)."""
        try:
            with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as zf:
                namelist = zf.namelist()
                dangerous = [
                    n for n in namelist
                    if "externalLinks" in n
                    or "xl/embeddings/" in n
                    or "oleObject" in n
                ]
                if dangerous:
                    msg = get_message("external_links")
                    raise FileSafetyError(
                        problem=msg["problem"],
                        corrective_step=msg["corrective_step"],
                        check_name="external_links",
                    )
        except zipfile.BadZipFile:
            pass

    def _check_formulas(self, file_bytes: bytes) -> None:
        """Check for value formulas (Requirement 4.8)."""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=True)
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.data_type == "f":
                            location = f"{ws.title}!{cell.coordinate}"
                            msg = get_message("formula_detected", location=location)
                            raise FileSafetyError(
                                problem=msg["problem"],
                                corrective_step=msg["corrective_step"],
                                location=location,
                                check_name="formula",
                            )
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            location = f"{ws.title}!{cell.coordinate}"
                            msg = get_message("formula_detected", location=location)
                            raise FileSafetyError(
                                problem=msg["problem"],
                                corrective_step=msg["corrective_step"],
                                location=location,
                                check_name="formula",
                            )
            wb.close()
        except FileSafetyError:
            raise
        except Exception:
            pass  # If we can't parse, it will be caught elsewhere
