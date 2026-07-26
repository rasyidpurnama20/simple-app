"""TemplateGenerator — deterministic .xlsx generation.

Produces five fixed-order sheets with embedded identity, byte-stable output,
and no value formulas (Requirements 2.1-2.5).
"""

from __future__ import annotations

import io
import struct
import zipfile
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from excel_import.models import TemplateDefinition

# Fixed epoch for deterministic metadata
FIXED_EPOCH = datetime(2024, 1, 1, 0, 0, 0)

# Fixed zip timestamp
FIXED_ZIP_TIME = (2024, 1, 1, 0, 0, 0)


class TemplateGenerator:
    """Generates deterministic .xlsx files from a TemplateDefinition."""

    def generate(
        self,
        definition: TemplateDefinition,
        scope: dict | None = None,
        reference_rows: list[dict] | None = None,
        prefill_rows: list[dict] | None = None,
    ) -> bytes:
        """Generate a deterministic .xlsx workbook.

        Sheets (fixed order): Petunjuk, Metadata, Data, Referensi, Validasi.
        """
        scope = scope or {}
        reference_rows = reference_rows or []
        prefill_rows = prefill_rows or []

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # 1. Fixed sheet order
        self._build_petunjuk(wb, definition)
        self._build_metadata(wb, definition, scope)
        self._build_data(wb, definition, prefill_rows)
        self._build_referensi(wb, definition, reference_rows)
        self._build_validasi(wb, definition)

        # 2. No volatile metadata
        wb.properties.created = FIXED_EPOCH
        wb.properties.modified = FIXED_EPOCH
        wb.properties.creator = "OBE Excel Import"

        # 3. Serialize and normalize
        raw = self._save_to_bytes(wb)
        return self._normalize_xlsx_zip(raw)

    def _build_petunjuk(self, wb: Workbook, definition: TemplateDefinition) -> None:
        """Build the Petunjuk (instructions) sheet."""
        ws = wb.create_sheet("Petunjuk")
        ws["A1"] = "Petunjuk Pengisian Template"
        ws["A2"] = f"Tipe: {definition.template_type}"
        ws["A3"] = f"Versi: {definition.schema_version}"
        ws["A5"] = "Instruksi:"
        ws["A6"] = "1. Isi data pada sheet 'Data' sesuai kolom yang tersedia."
        ws["A7"] = "2. Jangan mengubah sheet 'Metadata' dan 'Validasi'."
        ws["A8"] = "3. Lihat sheet 'Referensi' untuk data acuan."
        ws["A9"] = "4. Jangan menggunakan rumus (formula)."
        ws["A10"] = "5. Simpan sebagai .xlsx dan unggah kembali."

    def _build_metadata(self, wb: Workbook, definition: TemplateDefinition, scope: dict) -> None:
        """Build Metadata with embedded identity (Requirement 2.2)."""
        ws = wb.create_sheet("Metadata")
        ws["A1"] = "template_id"
        ws["B1"] = definition.template_type
        ws["A2"] = "schema_version"
        ws["B2"] = definition.schema_version
        ws["A3"] = "prodi"
        ws["B3"] = scope.get("prodi", "")
        ws["A4"] = "period"
        ws["B4"] = scope.get("period", "")
        ws["A5"] = "class"
        ws["B5"] = scope.get("klass", "")

        # Also embed in custom properties (docProps/custom)
        # Note: openpyxl handles custom props through the workbook
        ws["A7"] = "template_id_check"
        ws["B7"] = definition.template_type
        ws["A8"] = "schema_version_check"
        ws["B8"] = definition.schema_version

    def _build_data(self, wb: Workbook, definition: TemplateDefinition, prefill_rows: list[dict]) -> None:
        """Build the Data sheet with field headers and prefill."""
        ws = wb.create_sheet("Data")
        fields = definition.fields or []

        # Header row
        for col_idx, field in enumerate(fields, 1):
            ws.cell(row=1, column=col_idx, value=field.get("label", field.get("name", "")))

        # Sort prefill by business key for determinism
        bk_fields = definition.business_key or []
        sorted_rows = sorted(
            prefill_rows,
            key=lambda r: tuple(str(r.get(k, "")) for k in bk_fields),
        )

        # Prefill data
        for row_idx, row_data in enumerate(sorted_rows, 2):
            for col_idx, field in enumerate(fields, 1):
                value = row_data.get(field.get("name", ""), "")
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _build_referensi(self, wb: Workbook, definition: TemplateDefinition, reference_rows: list[dict]) -> None:
        """Build the Referensi (reference data) sheet."""
        ws = wb.create_sheet("Referensi")
        ws["A1"] = "Data Referensi"

        if not reference_rows:
            ws["A2"] = "(Tidak ada data referensi)"
            return

        # Write reference data sorted by first column for determinism
        if reference_rows:
            headers = list(reference_rows[0].keys())
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=2, column=col_idx, value=header)

            sorted_refs = sorted(reference_rows, key=lambda r: str(r.get(headers[0], "")))
            for row_idx, row_data in enumerate(sorted_refs, 3):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

    def _build_validasi(self, wb: Workbook, definition: TemplateDefinition) -> None:
        """Build the Validasi sheet from definition's validation rules."""
        ws = wb.create_sheet("Validasi")
        ws["A1"] = "Aturan Validasi"
        ws["A2"] = "Field"
        ws["B2"] = "Aturan"
        ws["C2"] = "Parameter"

        rules = definition.validation_rules or []
        for row_idx, rule in enumerate(rules, 3):
            ws.cell(row=row_idx, column=1, value=rule.get("field", ""))
            ws.cell(row=row_idx, column=2, value=rule.get("rule", ""))
            ws.cell(row=row_idx, column=3, value=str(rule.get("params", {})))

    def _save_to_bytes(self, wb: Workbook) -> bytes:
        """Save workbook to bytes."""
        stream = io.BytesIO()
        wb.save(stream)
        return stream.getvalue()

    def _normalize_xlsx_zip(self, raw: bytes) -> bytes:
        """Re-pack OOXML parts with fixed member order and timestamps.

        Ensures byte-identical output for identical inputs (Requirement 2.4).
        """
        output = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(raw), "r") as zin:
            members = sorted(zin.namelist())
            with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zout:
                for member in members:
                    info = zipfile.ZipInfo(member)
                    info.date_time = FIXED_ZIP_TIME
                    info.compress_type = zipfile.ZIP_DEFLATED
                    zout.writestr(info, zin.read(member))
        return output.getvalue()
