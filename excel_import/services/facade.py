"""ExcelImportService — transport-agnostic facade.

Composes all components and owns the lifecycle state machine.
Requirements: 7.2-7.7, 9.2.
"""

from __future__ import annotations

import hashlib
import io

from django.db import transaction

from excel_import.errors import DomainError, SchemaVersionMismatchError
from excel_import.models import ImportBatch, ImportStatus

from .commit_engine import CommitEngine, ReconciliationSummary
from .dry_run import DryRunReport, DryRunValidator
from .file_validator import FileValidator
from .scope_resolver import ImportScope, ScopeResolver
from .staging import ParsedRow, StagingArea
from .template_generator import TemplateGenerator
from .template_registry import TemplateRegistry


class ExcelImportService:
    """Facade composing registry, generator, validator, staging, dry-run, commit.

    All methods are transport-agnostic: accept plain bytes/ids/DTOs, return
    plain DTOs (Requirement 9.2). No HttpRequest/HttpResponse dependency.
    """

    def __init__(
        self,
        registry: TemplateRegistry | None = None,
        generator: TemplateGenerator | None = None,
        scope_resolver: ScopeResolver | None = None,
        file_validator: FileValidator | None = None,
        staging: StagingArea | None = None,
        dry_run: DryRunValidator | None = None,
        commit_engine: CommitEngine | None = None,
    ):
        self.registry = registry or TemplateRegistry()
        self.generator = generator or TemplateGenerator()
        self.scope_resolver = scope_resolver or ScopeResolver()
        self.file_validator = file_validator or FileValidator()
        self.staging = staging or StagingArea()
        self.dry_run = dry_run or DryRunValidator(registry=self.registry, staging=self.staging)
        self.commit_engine = commit_engine or CommitEngine(registry=self.registry, staging=self.staging)

    def generate_workbook(
        self,
        template_type: str,
        *,
        timeline_task_id: str | None = None,
        actor=None,
    ) -> bytes:
        """Generate a deterministic template workbook.

        Returns .xlsx bytes ready for download.
        """
        self.registry.require_implemented(template_type)
        definition = self.registry.get_current(template_type)
        scope, resolved_type = self.scope_resolver.resolve(template_type, timeline_task_id)
        reference_rows = self.scope_resolver.reference_data(definition, scope)
        prefill_rows = self.scope_resolver.prior_data(definition, scope)

        return self.generator.generate(
            definition=definition,
            scope={"prodi": scope.prodi, "period": scope.period, "klass": scope.klass},
            reference_rows=reference_rows,
            prefill_rows=prefill_rows,
        )

    @transaction.atomic
    def upload_and_dry_run(
        self,
        file_bytes: bytes,
        declared_mime: str = "",
        filename: str = "",
        template_type: str = "",
        schema_version: str = "",
        scope: dict | None = None,
        actor=None,
    ) -> DryRunReport:
        """Upload, validate, stage, and dry-run. No owning-module writes.

        Lifecycle: creates batch as Diunggah, sets Divalidasi on pass or
        Ditolak on rejection (Requirements 7.3, 7.4).
        """
        from core.models import DemoUser, ProgramOfStudy

        # 1. File validation (rejects unsafe files)
        self.file_validator.validate_file(file_bytes, declared_mime, filename)

        # 2. Parse embedded identity from the file
        parsed_type, parsed_version = self._read_embedded_identity(file_bytes)
        template_type = template_type or parsed_type
        schema_version = schema_version or parsed_version

        # 3. Create batch
        scope = scope or {}
        actor_obj = None
        prodi_obj = None
        if actor:
            if isinstance(actor, int):
                actor_obj = DemoUser.objects.filter(pk=actor).first()
            else:
                actor_obj = actor
        if actor_obj and actor_obj.prodi:
            prodi_obj = actor_obj.prodi
        else:
            prodi_obj = ProgramOfStudy.objects.first()

        content_hash = hashlib.sha256(file_bytes).hexdigest()

        batch = ImportBatch.objects.create(
            template_type=template_type,
            schema_version=schema_version,
            import_prodi=scope.get("prodi", ""),
            period=scope.get("period", ""),
            klass=scope.get("klass", ""),
            batch_status=ImportStatus.DIUNGGAH,
            content_hash=content_hash,
            prodi=prodi_obj,
            owner=actor_obj,
            creator=actor_obj,
            status="active",
        )

        # 4. Parse Data sheet and stage rows
        parsed_rows = self._parse_data_sheet(file_bytes, template_type, schema_version)
        self.staging.stage(batch, parsed_rows)

        # 5. Dry-run validation
        try:
            report = self.dry_run.dry_run(batch)
            batch.batch_status = ImportStatus.DIVALIDASI
            batch.save(update_fields=["batch_status", "modified_time"])
            return report
        except (SchemaVersionMismatchError, DomainError):
            batch.batch_status = ImportStatus.DITOLAK
            batch.save(update_fields=["batch_status", "modified_time"])
            raise

    def commit(self, batch_id: str, actor=None) -> ReconciliationSummary:
        """Commit a validated batch atomically (Requirement 6.1).

        Lifecycle: Dikomit on success, Digagalkan on failure (7.5, 7.6).
        """
        batch = ImportBatch.objects.get(pk=batch_id)
        if batch.batch_status != ImportStatus.DIVALIDASI:
            raise DomainError(
                problem="Batch belum divalidasi atau sudah diproses.",
                corrective_step="Pastikan batch dalam status Divalidasi sebelum commit.",
            )
        return self.commit_engine.commit(batch)

    def get_batch(self, batch_id: str):
        """Get a batch by ID."""
        return ImportBatch.objects.get(pk=batch_id)

    def get_report(self, batch_id: str) -> DryRunReport | None:
        """Reconstruct a dry-run report for a batch."""
        from .dry_run import DryRunReportEntry
        batch = ImportBatch.objects.get(pk=batch_id)
        rows = list(batch.rows.all().order_by("row_index"))
        report = DryRunReport(batch_id=str(batch.id), total_rows=len(rows))
        for row in rows:
            report.entries.append(DryRunReportEntry(
                row_index=row.row_index,
                business_key=row.business_key,
                classification=row.classification or "",
                cell_errors=row.cell_errors or [],
            ))
        return report

    def _read_embedded_identity(self, file_bytes: bytes) -> tuple[str, str]:
        """Read Template_Id and Schema_Version from Metadata sheet."""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            if "Metadata" in wb.sheetnames:
                ws = wb["Metadata"]
                template_id = str(ws["B1"].value or "")
                schema_version = str(ws["B2"].value or "")
                wb.close()
                return template_id, schema_version
            wb.close()
        except Exception:
            pass
        return "", ""

    def _parse_data_sheet(self, file_bytes: bytes, template_type: str, schema_version: str) -> list[ParsedRow]:
        """Parse the Data sheet into ParsedRow objects."""
        parsed_rows = []
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            if "Data" not in wb.sheetnames:
                wb.close()
                return parsed_rows

            ws = wb["Data"]
            rows_iter = ws.iter_rows()

            # First row is header
            header_row = next(rows_iter, None)
            if header_row is None:
                wb.close()
                return parsed_rows

            headers = [str(cell.value or "").strip() for cell in header_row]

            # Get business key fields from definition
            defn = None
            if self.registry:
                defn = self.registry.get_version(template_type, schema_version)
            bk_fields = defn.business_key if defn else headers[:1]

            for row_idx, row in enumerate(rows_iter, 2):
                values = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers):
                        values[headers[col_idx]] = cell.value if cell.value is not None else ""

                # Skip completely empty rows
                if not any(str(v).strip() for v in values.values()):
                    continue

                # Compute business key
                bk_parts = [str(values.get(k, "")).strip() for k in bk_fields]
                business_key = "|".join(bk_parts)

                parsed_rows.append(ParsedRow(
                    row_index=row_idx,
                    raw_values=values,
                    business_key=business_key,
                ))

            wb.close()
        except Exception:
            pass
        return parsed_rows
